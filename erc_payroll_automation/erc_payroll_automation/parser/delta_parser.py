"""Delta-mode parser.

For projects whose customer file carries only *deltas* (variable earnings /
deductions) rather than the full salary breakdown — currently Datavolt and
Airproducts. Base salary comes from system Employee data; the file only adjusts
overtime / other income / deductions.

Output is a normal set of Payroll Import Row records, so the existing
Generate Outputs step renders the standard Elite internal sheet unchanged.

Population differs by mode (per the client):
    Datavolt    — ALL active project employees. Match deltas by Employee ID
                  (fallback employee_id_from_client). The few employees in
                  'Variable Earnings' / 'Variable Deductions' get adjusted;
                  everyone else is pure system data.
                  Variable Earnings.Amount   -> Other Income
                  Variable Deductions.Amount -> Deductions (abs)
    Airproducts — Employees are SPECIFIED BY the file (matched on IQAMA). Each
                  file row -> one employee, base from system + the file deltas.
                  Overtime Amount            -> Overtime
                  other allowance columns sum-> Other Income
                  DEDUCTION AMOUNT           -> Deductions (abs)

Scoping note: we deliberately do NOT filter by Work Location here. These are
single-project clients and Employee.custom_location is often blank, which would
wrongly drop almost everyone. Airproducts also has a global IQAMA fallback so an
existing active employee still matches even if their project field is off.
"""

import json

import frappe
from frappe import _
from frappe.utils import getdate
from openpyxl import load_workbook

from ..parser_constants import NUMERIC_FIELDS
from .file_parser import _resolve_attached_file_path, _trunc, _json_safe


DELTA_DATAVOLT = "Deltas (Datavolt)"
DELTA_AIRPRODUCTS = "Deltas (Airproducts)"

EMP_FIELDS = [
    "name", "employee_name", "employment_type", "date_of_joining",
    "nationality", "added_to_gosi", "bank_name", "bank_ac_no",
    "iqama_national_id", "passport_number", "employee_id_from_client", "status",
]


def is_delta_mode(template):
    mode = (getattr(template, "input_mode", None) or "Full from File")
    return mode.startswith("Deltas")


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

def run_delta_parse(run, template):
    """Build Payroll Import Rows (base from system + file deltas) and persist.
    Sets status -> Reconciliation Pending."""
    mode = template.input_mode
    if mode == DELTA_DATAVOLT:
        rows, unmatched = _datavolt_rows(run, template)
    elif mode == DELTA_AIRPRODUCTS:
        rows, unmatched = _airproducts_rows(run, template)
    else:
        raise ValueError(f"Unsupported delta input_mode: {mode!r}")

    if not rows:
        raise ValueError(_(
            "No employee rows produced for {0}. Check that the project's "
            "employees exist in the system and the file matches them."
        ).format(mode))

    _persist(run, rows, unmatched)

    frappe.log_error(
        title=f"Payroll Import delta parse COMPLETE {run.name}",
        message=(f"mode={mode} rows={len(rows)} "
                 f"unmatched_file_rows={len(unmatched)}"),
    )


# ---------------------------------------------------------------------------
# Datavolt: population = all active project employees, deltas overlaid
# ---------------------------------------------------------------------------

def _datavolt_rows(run, template):
    employees = _scope_employees(template)
    if not employees:
        raise ValueError(_(
            "No active employees found for project {0}."
        ).format(template.project))

    by_name = {str(e.name).strip(): e.name for e in employees}
    by_client = {
        str(e.employee_id_from_client).strip(): e.name
        for e in employees if e.get("employee_id_from_client")
    }

    deltas, unmatched = {}, []
    path = _resolve_attached_file_path(run.source_file)
    wb = load_workbook(path, data_only=True, read_only=True)
    try:
        for raw_key, amount in _datavolt_sheet(wb, "Variable Earnings"):
            emp = by_name.get(_norm_id(raw_key)) or by_client.get(_norm_id(raw_key))
            if not emp:
                unmatched.append({"sheet": "Variable Earnings",
                                  "raw_key": raw_key, "amount": amount})
                continue
            deltas.setdefault(emp, _zero())["other_income"] += amount
        for raw_key, amount in _datavolt_sheet(wb, "Variable Deductions"):
            emp = by_name.get(_norm_id(raw_key)) or by_client.get(_norm_id(raw_key))
            if not emp:
                unmatched.append({"sheet": "Variable Deductions",
                                  "raw_key": raw_key, "amount": amount})
                continue
            deltas.setdefault(emp, _zero())["deductions"] += abs(amount)
    finally:
        wb.close()

    period_days = _period_days(run)
    rows = [
        _build_row(e, _base_amounts(e, template),
                   deltas.get(e.name) or _zero(), period_days)
        for e in employees
    ]
    return rows, unmatched


def _datavolt_sheet(wb, sheet_name):
    """Yield (raw_employee_code, amount) for a Datavolt delta sheet.
    Header is the first row (within the top 8) containing 'employee code'."""
    if sheet_name not in wb.sheetnames:
        return []
    rows = list(wb[sheet_name].iter_rows(values_only=True))
    hidx, norm = _find_header(rows, ["employee code"])
    if hidx is None:
        return []
    key_c = _col(norm, "employee code")
    amt_c = _col(norm, "amount")
    out = []
    for r in rows[hidx + 1:]:
        if key_c is None or key_c >= len(r):
            continue
        raw_key = r[key_c]
        if raw_key in (None, ""):
            continue
        out.append((raw_key, _to_float(_cell(r, amt_c))))
    return out


# ---------------------------------------------------------------------------
# Airproducts: population = file rows, matched to employees by IQAMA
# ---------------------------------------------------------------------------

def _airproducts_rows(run, template):
    file_rows = _read_airproducts_file(run)
    if not file_rows:
        raise ValueError(_("No data rows found in the Airproducts file "
                           "(could not locate the IQAMA column)."))

    # Match against project employees first; fall back to a global IQAMA lookup
    # so an existing active employee still matches even with a wrong project.
    proj_emps = _scope_employees(template)
    by_iqama = {
        _norm_id(e.iqama_national_id): e
        for e in proj_emps if e.get("iqama_national_id")
    }
    missing = [fr for fr in file_rows if _norm_id(fr["iqama"]) not in by_iqama]
    if missing:
        for e in _employees_by_iqama([_norm_id(fr["iqama"]) for fr in missing],
                                     template):
            by_iqama.setdefault(_norm_id(e.iqama_national_id), e)

    period_days = _period_days(run)
    rows, unmatched, seen = [], [], {}
    for fr in file_rows:
        iq = _norm_id(fr["iqama"])
        e = by_iqama.get(iq)
        if not e:
            unmatched.append({"sheet": "Allowances", "raw_key": fr["iqama"],
                              "amount": fr["overtime"] + fr["other_income"] - fr["deductions"]})
            continue
        d = {"overtime": fr["overtime"], "other_income": fr["other_income"],
             "deductions": fr["deductions"]}
        if e.name in seen:
            # Same employee appears twice in the file — accumulate the deltas.
            row = seen[e.name]
            row["overtime"] += d["overtime"]
            row["other_income"] += d["other_income"]
            row["deductions"] += d["deductions"]
            row["net_salary_from_file"] = _net(row)
            continue
        row = _build_row(e, _base_amounts(e, template), d, period_days)
        rows.append(row)
        seen[e.name] = row
    return rows, unmatched


def _read_airproducts_file(run):
    """Return list of {iqama, overtime, other_income, deductions} from the file."""
    path = _resolve_attached_file_path(run.source_file)
    wb = load_workbook(path, data_only=True, read_only=True)
    try:
        sheet = _airproducts_sheet(wb)
        if not sheet:
            return []
        ws_rows, hidx, norm = sheet
        iq_c = _col(norm, "iqama")
        ot_c = _col(norm, "overtime amount", "overtime")
        ded_c = _col(norm, "deduction amount", "deduction")
        add_cs = [
            _col(norm, h) for h in
            ("other allowance", "utility allowance", "trip allowance",
             "ticket allowance", "phone allowance")
        ]
        out = []
        for r in ws_rows[hidx + 1:]:
            if iq_c is None or iq_c >= len(r):
                continue
            raw_iq = r[iq_c]
            if raw_iq in (None, ""):
                continue
            out.append({
                "iqama": raw_iq,
                "overtime": _to_float(_cell(r, ot_c)),
                "other_income": sum(_to_float(_cell(r, c)) for c in add_cs if c is not None),
                "deductions": abs(_to_float(_cell(r, ded_c))),
            })
        return out
    finally:
        wb.close()


def _airproducts_sheet(wb):
    """Find the data sheet (header containing 'IQAMA'). Returns (rows, hidx, norm)."""
    for sn in wb.sheetnames:
        rows = list(wb[sn].iter_rows(values_only=True))
        hidx, norm = _find_header(rows, ["iqama"])
        if hidx is not None:
            return rows, hidx, norm
    return None


# ---------------------------------------------------------------------------
# Employee scope + base salary
# ---------------------------------------------------------------------------

def _scope_employees(template):
    """Active employees on the template's project. No Work Location filter —
    custom_location is often blank for these clients and would drop everyone."""
    filters = {"status": "Active"}
    if template.project:
        filters["project"] = template.project

    fields = EMP_FIELDS[:]
    for cf in _base_fieldnames(template):
        if cf not in fields and _field_exists("Employee", cf):
            fields.append(cf)
    return frappe.get_all("Employee", filters=filters, fields=fields)


def _employees_by_iqama(iqamas, template):
    """Global fallback: fetch active employees by IQAMA, any project."""
    iqamas = [i for i in dict.fromkeys(iqamas) if i]
    if not iqamas:
        return []
    fields = EMP_FIELDS[:]
    for cf in _base_fieldnames(template):
        if cf not in fields and _field_exists("Employee", cf):
            fields.append(cf)
    return frappe.get_all(
        "Employee",
        filters={"status": "Active", "iqama_national_id": ["in", iqamas]},
        fields=fields,
    )


def _base_fieldnames(template):
    return [
        template.no_file_emp_field_basic or "basic_salary",
        template.no_file_emp_field_housing or "housing_allowance",
        template.no_file_emp_field_transportation or "transport_allowance",
        template.no_file_emp_field_other_allowance or "food_allowance",
    ]


def _base_amounts(emp, template):
    fb, fh, ft, fo = _base_fieldnames(template)
    return {
        "basic": _to_float(emp.get(fb)),
        "housing": _to_float(emp.get(fh)),
        "transport": _to_float(emp.get(ft)),
        "other": _to_float(emp.get(fo)),
    }


def _field_exists(doctype, fieldname):
    try:
        return bool(frappe.get_meta(doctype).get_field(fieldname))
    except Exception:
        return False


# ---------------------------------------------------------------------------
# Header / cell helpers
# ---------------------------------------------------------------------------

def _zero():
    return {"overtime": 0.0, "other_income": 0.0, "deductions": 0.0}


def _norm(v):
    return " ".join(str(v).strip().lower().split()) if v is not None else ""


def _find_header(rows, must_have, max_scan=8):
    wanted = [_norm(m) for m in must_have]
    for i, r in enumerate(rows[:max_scan]):
        norm = [_norm(c) for c in r]
        if all(any(w == h or w in h for h in norm) for w in wanted):
            return i, norm
    return None, None


def _col(norm_headers, *candidates):
    for cand in candidates:
        c = _norm(cand)
        for idx, h in enumerate(norm_headers):
            if h == c:
                return idx
    for cand in candidates:
        c = _norm(cand)
        for idx, h in enumerate(norm_headers):
            if c and c in h:
                return idx
    return None


def _cell(row, idx):
    if idx is None or idx >= len(row):
        return None
    return row[idx]


def _norm_id(value):
    """Normalize a numeric-ish code: 10019 / 10019.0 / '10019 ' -> '10019'."""
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    if isinstance(value, int):
        return str(value)
    return str(value).strip()


def _to_float(value):
    if value in (None, ""):
        return 0.0
    if isinstance(value, bool):
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    s = str(value).strip().replace(",", "")
    try:
        return float(s)
    except ValueError:
        return 0.0


def _period_days(run):
    try:
        d = (getdate(run.payroll_period_end) - getdate(run.payroll_period_start)).days + 1
        return d if d > 0 else 30
    except Exception:
        return 30


# ---------------------------------------------------------------------------
# Build + persist rows
# ---------------------------------------------------------------------------

def _net(row):
    return (row["total_per_contract"] + row["overtime"]
            + row["other_income"] - row["deductions"])


def _build_row(emp, base, d, period_days):
    total = base["basic"] + base["housing"] + base["transport"] + base["other"]
    row = {
        "employee": emp.name,
        "match_method": "system_base",
        "match_confidence": 1.0,
        "raw_id_value": emp.get("iqama_national_id") or emp.get("passport_number") or "",
        "raw_name": emp.get("employee_name") or "",
        "raw_iban": emp.get("bank_ac_no") or "",
        "raw_nationality": emp.get("nationality") or "",
        "basic_per_contract": base["basic"],
        "housing_per_contract": base["housing"],
        "transportation_per_contract": base["transport"],
        "other_allowance_per_contract": base["other"],
        "total_per_contract": total,
        "working_days": period_days,
        "basic_per_working_days": base["basic"],
        "housing_per_working_days": base["housing"],
        "transportation_per_working_days": base["transport"],
        "other_allowance_per_working_days": base["other"],
        "overtime": d["overtime"],
        "other_income": d["other_income"],
        "deductions": d["deductions"],
        "hq_deductions": 0.0,
        "gosi_from_file": None,
    }
    row["net_salary_from_file"] = _net(row)
    return row


def _persist(run, rows, unmatched):
    frappe.db.delete("Payroll Import Row", {"parent": run.name})
    frappe.db.delete("Payroll Unmatched Source Row", {"parent": run.name})
    frappe.db.delete("Payroll Unaccounted Employee", {"parent": run.name})

    for i, p in enumerate(rows, start=1):
        doc = frappe.new_doc("Payroll Import Row")
        doc.parent = run.name
        doc.parenttype = "Payroll Import Run"
        doc.parentfield = "parsed_rows"
        doc.idx = i
        doc.row_number_in_file = 0
        doc.employee = p["employee"]
        doc.match_method = p["match_method"]
        doc.match_confidence = p["match_confidence"]
        doc.raw_id_value = _trunc(p.get("raw_id_value"), 140)
        doc.raw_name = _trunc(p.get("raw_name"), 140)
        doc.raw_iban = _trunc(p.get("raw_iban"), 140)
        doc.raw_nationality = _trunc(p.get("raw_nationality"), 140)
        for fname in NUMERIC_FIELDS:
            v = p.get(fname)
            if v is not None:
                setattr(doc, fname, v)
        doc.raw_data_json = json.dumps(
            {k: _json_safe(v) for k, v in p.items()}, default=str)
        doc.validation_status = "OK"
        doc.validation_messages = ""
        doc.flags.ignore_permissions = True
        doc.flags.ignore_mandatory = True
        doc.db_insert()

    for i, u in enumerate(unmatched, start=1):
        doc = frappe.new_doc("Payroll Unmatched Source Row")
        doc.parent = run.name
        doc.parenttype = "Payroll Import Run"
        doc.parentfield = "unmatched_source_rows"
        doc.idx = i
        doc.row_number_in_file = 0
        doc.raw_id_value = _trunc(str(u.get("raw_key")), 140)
        doc.raw_name = ""
        doc.reason_unmatched = f"code/iqama not found in system ({u.get('sheet')})"
        doc.raw_data_json = json.dumps(
            {k: _json_safe(v) for k, v in u.items()}, default=str)
        doc.flags.ignore_permissions = True
        doc.flags.ignore_mandatory = True
        doc.db_insert()

    total = len(rows)
    frappe.db.set_value("Payroll Import Run", run.name, {
        "rows_matched": total,
        "rows_unmatched_in_file": len(unmatched),
        "rows_unaccounted_in_system": 0,
        "rows_with_warnings": 0,
        "rows_with_errors": 0,
        "validation_pass_rate": 100.0 if total else 0.0,
        "status": "Reconciliation Pending",
        "parse_error_log": "",
    }, update_modified=False)
    frappe.db.commit()
