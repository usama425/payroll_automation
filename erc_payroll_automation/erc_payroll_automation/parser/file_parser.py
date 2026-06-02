"""file_parser — opens the customer Excel, applies the template's column map,
runs employee matching + reconciliation, and persists results onto the Run.

Entry point: run_parse(run_name, user) — called from PayrollImportRun.trigger_parse via frappe.enqueue.
"""

import json
import os

import frappe
from frappe import _
from openpyxl import load_workbook

from . import transforms
from . import employee_matcher
from . import reconciliation
from . import validators
from ..parser_constants import NUMERIC_FIELDS  # re-exported below for back-compat


__all__ = ["NUMERIC_FIELDS", "run_parse"]


def run_parse(run_name, user=None):
    """Background-job entry point. Called from PayrollImportRun.trigger_parse via frappe.enqueue."""
    frappe.log_error(
        title=f"Payroll Import parse STARTED: {run_name}",
        message=f"Worker picked up parse job for {run_name} (user={user})",
    )

    run = frappe.get_doc("Payroll Import Run", run_name)
    template = frappe.get_doc("Payroll Import Template", run.template)

    try:
        # Delta-mode projects (Datavolt, Airproducts): base salary from system,
        # file supplies only earnings/deductions. Different parse path entirely.
        from . import delta_parser
        if delta_parser.is_delta_mode(template):
            delta_parser.run_delta_parse(run, template)
            frappe.db.commit()
            return

        file_path = _resolve_attached_file_path(run.source_file)
        wb = load_workbook(file_path, data_only=True, read_only=True)
        sheet_name = template.sheet_name_override or _pick_first_data_sheet(wb)
        if sheet_name not in wb.sheetnames:
            raise ValueError(
                f"Sheet '{sheet_name}' not found in source file. "
                f"Available sheets: {wb.sheetnames}"
            )
        ws = wb[sheet_name]
        run.db_set("source_file_sheet_used", sheet_name)

        header_row_idx = int(template.header_row_index or 1)
        headers = _read_header_row(ws, header_row_idx)
        col_map = _build_column_index_map(template.column_map, headers)

        parsed_rows = []
        rows_read = 0
        for row_idx, row in enumerate(
            ws.iter_rows(min_row=header_row_idx + 1, values_only=True),
            start=header_row_idx + 1,
        ):
            if all(c in (None, "") for c in row):
                continue
            rows_read += 1
            parsed = _parse_row(row, col_map)
            parsed["_row_index"] = row_idx

            if template.skip_blank_id_rows and not parsed.get("raw_id_value"):
                continue
            parsed_rows.append(parsed)

        wb.close()
        run.db_set("source_file_rows_total", rows_read)
        frappe.log_error(
            title=f"Payroll Import parse: file read {run_name}",
            message=f"Read {rows_read} rows from sheet '{sheet_name}'. Now matching.",
        )

        # Build indexes
        indexes = employee_matcher.build_indexes(template)
        frappe.log_error(
            title=f"Payroll Import parse: indexes built {run_name}",
            message=(
                f"Built employee indexes for {len(indexes['employees'])} employees. "
                f"Matching {len(parsed_rows)} parsed rows."
            ),
        )

        # Match — wrap each row so one bad row doesn't kill the whole parse
        matched, unmatched = [], []
        for p in parsed_rows:
            try:
                m = employee_matcher.match_row(p, indexes, template)
            except Exception:
                frappe.log_error(
                    title=f"Payroll Import parse: match_row error row {p.get('_row_index')}",
                    message=frappe.get_traceback(),
                )
                m = {"employee": None, "method": None, "confidence": 0.0,
                     "reason": "match_error", "suggestions": []}
            p["_match"] = m
            (matched if m["employee"] else unmatched).append(p)

        frappe.log_error(
            title=f"Payroll Import parse: matching done {run_name}",
            message=f"matched={len(matched)} unmatched={len(unmatched)}. Computing unaccounted.",
        )

        # Reconcile: who's expected but absent?
        matched_emp_ids = {p["_match"]["employee"] for p in matched}
        expected_emp_ids = set(indexes["employees"].keys())
        unaccounted_emp_ids = expected_emp_ids - matched_emp_ids

        frappe.log_error(
            title=f"Payroll Import parse: reconcile done {run_name}",
            message=f"unaccounted={len(unaccounted_emp_ids)}. Pre-validating matched rows.",
        )

        # Run validators in-memory BEFORE persisting — avoids a second DB round-trip
        warnings, errors = _pre_validate_matched(matched, template, indexes)

        frappe.log_error(
            title=f"Payroll Import parse: pre-validate done {run_name}",
            message=f"warnings={warnings} errors={errors}. Persisting via db_insert.",
        )

        # Persist with direct db_insert() — avoids the slow run.save() with N children
        _persist_results(run, matched, unmatched, unaccounted_emp_ids, indexes, warnings, errors)

        frappe.log_error(
            title=f"Payroll Import parse: persist done {run_name}",
            message="All child rows saved. Setting status.",
        )

        run.db_set("status", "Reconciliation Pending")
        run.db_set("parse_error_log", "")
        frappe.db.commit()

        frappe.log_error(
            title=f"Payroll Import parse COMPLETE {run_name}",
            message=(
                f"Status=Reconciliation Pending. "
                f"matched={len(matched)} unmatched={len(unmatched)} "
                f"unaccounted={len(unaccounted_emp_ids)} "
                f"warnings={warnings} errors={errors}"
            ),
        )

    except Exception:
        frappe.log_error(
            title=f"Payroll Import parse failed: {run_name}",
            message=frappe.get_traceback(),
        )
        try:
            run.db_set("parse_error_log", frappe.get_traceback())
            run.db_set("status", "Draft")
        except Exception:
            pass
        raise


# ---------------------------------------------------------------------------
# Pre-validation helpers
# ---------------------------------------------------------------------------

class _RowProxy:
    """Lightweight proxy that mimics a Payroll Import Row doc for in-memory validation."""

    def __init__(self, p, employee):
        # Pre-initialise every numeric and string field to None so validators
        # can safely call getattr without AttributeError.
        for fname in NUMERIC_FIELDS:
            setattr(self, fname, None)
        self.raw_id_value = None
        self.raw_name = None
        self.raw_iban = None
        self.raw_nationality = None

        # Overlay actual parsed values
        for k, v in p.items():
            if not k.startswith("_"):
                setattr(self, k, v)

        self.employee = employee
        self.validation_status = "OK"
        self.validation_messages = None


def _pre_validate_matched(matched, template, indexes):
    """Run validators in memory on matched rows before DB persist.

    Writes ``_validation_status`` and ``_validation_messages`` back onto each
    dict in ``matched``.  Returns (warnings_count, errors_count).
    """

    class _FakeRun:
        def __init__(self, rows):
            self.parsed_rows = rows
            self.rows_with_warnings = 0
            self.rows_with_errors = 0

    proxies = [_RowProxy(p, p["_match"]["employee"]) for p in matched]
    fake_run = _FakeRun(proxies)
    try:
        validators.validate_run(fake_run, template, indexes)
    except Exception:
        frappe.log_error(
            title="Payroll Import parse: pre-validate error (non-fatal)",
            message=frappe.get_traceback(),
        )
        # Non-fatal — continue with all rows marked OK
        return 0, 0

    for p, proxy in zip(matched, proxies):
        p["_validation_status"] = proxy.validation_status
        p["_validation_messages"] = proxy.validation_messages or ""

    return fake_run.rows_with_warnings, fake_run.rows_with_errors


# ---------------------------------------------------------------------------
# Persist helpers — use db_insert() to bypass the slow run.save() pattern
# ---------------------------------------------------------------------------

def _persist_results(run, matched, unmatched, unaccounted_emp_ids, indexes,
                     warnings=0, errors=0):
    """Persist child rows using individual db_insert() calls.

    Avoids the slow Document.save() path that batches 76+ rows in one ORM
    cycle and was causing timeouts on Frappe Cloud workers.
    """
    # Clear previous data from DB
    frappe.db.delete("Payroll Import Row", {"parent": run.name})
    frappe.db.delete("Payroll Unmatched Source Row", {"parent": run.name})
    frappe.db.delete("Payroll Unaccounted Employee", {"parent": run.name})

    # Insert matched rows
    for i, p in enumerate(matched, start=1):
        doc = frappe.new_doc("Payroll Import Row")
        doc.parent = run.name
        doc.parenttype = "Payroll Import Run"
        doc.parentfield = "parsed_rows"
        doc.idx = i
        doc.row_number_in_file = p["_row_index"]
        doc.employee = p["_match"]["employee"]
        doc.match_method = p["_match"]["method"]
        doc.match_confidence = p["_match"]["confidence"]
        doc.raw_id_value = _trunc(p.get("raw_id_value"), 140)
        doc.raw_name = _trunc(p.get("raw_name"), 140)
        doc.raw_iban = _trunc(p.get("raw_iban"), 140)
        doc.raw_nationality = _trunc(p.get("raw_nationality"), 140)
        for fname in NUMERIC_FIELDS:
            v = p.get(fname)
            if v is not None:
                setattr(doc, fname, v)
        doc.raw_data_json = json.dumps(
            {k: _json_safe(v) for k, v in p.items() if not k.startswith("_")},
            default=str,
        )
        doc.validation_status = p.get("_validation_status", "OK")
        doc.validation_messages = p.get("_validation_messages") or ""
        doc.flags.ignore_permissions = True
        doc.flags.ignore_mandatory = True
        doc.db_insert()

    # Insert unmatched rows
    for i, p in enumerate(unmatched, start=1):
        doc = frappe.new_doc("Payroll Unmatched Source Row")
        doc.parent = run.name
        doc.parenttype = "Payroll Import Run"
        doc.parentfield = "unmatched_source_rows"
        doc.idx = i
        doc.row_number_in_file = p["_row_index"]
        doc.raw_name = _trunc(p.get("raw_name"), 140)
        doc.raw_id_value = _trunc(p.get("raw_id_value"), 140)
        doc.raw_iban = _trunc(p.get("raw_iban"), 140)
        doc.raw_nationality = _trunc(p.get("raw_nationality"), 140)
        doc.reason_unmatched = p["_match"].get("reason", "no_match")
        suggestions = p["_match"].get("suggestions") or []
        if suggestions:
            doc.suggested_employees = "\n".join(
                f"{s['employee']} - {s['name']} (score: {s['score']})"
                for s in suggestions[:3]
            )
        doc.raw_data_json = json.dumps(
            {k: _json_safe(v) for k, v in p.items() if not k.startswith("_")},
            default=str,
        )
        doc.flags.ignore_permissions = True
        doc.flags.ignore_mandatory = True
        doc.db_insert()

    # Insert unaccounted employees
    for i, u in enumerate(
        reconciliation.classify_unaccounted(unaccounted_emp_ids, indexes, run), start=1
    ):
        doc = frappe.new_doc("Payroll Unaccounted Employee")
        doc.parent = run.name
        doc.parenttype = "Payroll Import Run"
        doc.parentfield = "unaccounted_employees"
        doc.idx = i
        doc.employee = u["employee"]
        doc.auto_reason = u.get("auto_reason") or ""
        if u.get("auto_category"):
            doc.category = u["auto_category"]
        doc.flags.ignore_permissions = True
        doc.flags.ignore_mandatory = True
        doc.db_insert()

    # Update run counters directly — no run.save() needed
    total = len(matched)
    pass_rate = round(((total - errors) / total) * 100, 2) if total > 0 else 0.0
    frappe.db.set_value("Payroll Import Run", run.name, {
        "rows_matched": total,
        "rows_unmatched_in_file": len(unmatched),
        "rows_unaccounted_in_system": len(unaccounted_emp_ids),
        "rows_with_warnings": warnings,
        "rows_with_errors": errors,
        "validation_pass_rate": pass_rate,
    }, update_modified=False)
    frappe.db.commit()


# ---------------------------------------------------------------------------
# File / header helpers
# ---------------------------------------------------------------------------

def _resolve_attached_file_path(file_url):
    """Convert a Frappe File URL (/files/foo.xlsx or /private/files/foo.xlsx)
    into an absolute filesystem path. Also tolerates absolute paths."""
    if not file_url:
        raise ValueError(_("Source file is empty"))
    if file_url.startswith("/files/"):
        return os.path.join(frappe.get_site_path("public"), file_url.lstrip("/"))
    if file_url.startswith("/private/files/"):
        return os.path.join(frappe.get_site_path(), file_url.lstrip("/"))
    if file_url.startswith("/"):
        return os.path.join(frappe.get_site_path(), file_url.lstrip("/"))
    return file_url


def _pick_first_data_sheet(wb):
    """Pick the first sheet with at least 2 non-empty rows in its first 10 rows."""
    for sn in wb.sheetnames:
        ws = wb[sn]
        non_empty = 0
        for row in ws.iter_rows(min_row=1, max_row=10, values_only=True):
            if any(c not in (None, "") for c in row):
                non_empty += 1
                if non_empty >= 2:
                    return sn
    return wb.sheetnames[0]


def _read_header_row(ws, row_idx):
    return list(next(ws.iter_rows(min_row=row_idx, max_row=row_idx, values_only=True)))


def _normalize_header(h):
    return str(h).strip().lower() if h is not None else ""


def _build_column_index_map(column_map_rows, headers):
    """Returns {target_field: {"index": 0-based int or None, "transform": str, "required": bool}}."""
    normalized = [_normalize_header(h) for h in headers]
    result = {}
    for r in column_map_rows or []:
        if not r.target_field:
            continue
        idx = None
        if r.source_column_header:
            wanted = _normalize_header(r.source_column_header)
            try:
                idx = normalized.index(wanted)
            except ValueError:
                idx = None
        if idx is None and r.column_index_fallback:
            try:
                idx = int(r.column_index_fallback) - 1
            except (TypeError, ValueError):
                idx = None
        result[r.target_field] = {
            "index": idx,
            "transform": r.transform,
            "required": bool(r.is_required),
        }
    return result


def _parse_row(row, col_map):
    result = {}
    for target, spec in col_map.items():
        idx = spec["index"]
        if idx is None or idx >= len(row) or idx < 0:
            result[target] = None
            continue
        raw = row[idx]
        result[target] = transforms.apply(spec["transform"], raw)
    return result


def _trunc(value, n):
    if value is None:
        return None
    s = str(value)
    return s[:n]


def _json_safe(v):
    if v is None or isinstance(v, (str, int, float, bool)):
        return v
    return str(v)
