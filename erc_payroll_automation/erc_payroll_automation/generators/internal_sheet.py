"""Build the ERC internal Excel ("RSG-<Project> Payroll- <Month> <Year>.xlsx").

Layout (matches historical files like RSG- Misk School Payroll- February 2026.xlsx):
    Row 1: "Elite - RSP- <Project Name> Payroll" banner
    Row 2: "For the month of <Month> <Year>" banner
    Row 3: blank
    Row 4: column headers (33 columns)
    Row 5+: one row per matched employee, cells highlighted by validation status
"""

from datetime import datetime

import frappe
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

from ..parser.bank_codes import bank_display_for
from .file_attach import save_excel_and_link


# Column spec — (header, fieldname_on_payroll_import_row, kind)
# Kind: "data" reads from row, "emp" reads from Employee, "computed" calls function below.
COLUMNS = [
    ("S/N",                  None,                            "computed:serial"),
    ("ERP ID No.",           "employee",                      "data"),
    ("Name",                 None,                            "emp:employee_name"),
    ("Contract",             None,                            "emp:employment_type"),
    ("Company",              None,                            "computed:company_label"),
    ("Nationality",          "raw_nationality",               "data"),
    ("ID/Iqama/ passprot",   "raw_id_value",                  "data"),
    ("Bank Name",            None,                            "computed:bank_name"),
    ("IBAN/ Account number", "raw_iban",                      "computed:iban"),
    ("Basic per contract",   "basic_per_contract",            "data"),
    ("Working Days",         "working_days",                  "data"),
    ("Baisc",                "basic_per_working_days",        "data"),
    ("Housing",              "housing_per_working_days",      "data"),
    ("Transportaion",        "transportation_per_working_days","data"),
    ("Other Allowance",      "other_allowance_per_working_days","data"),
    ("Total Salary",         "total_per_contract",            "data"),
    ("Overtime",             "overtime",                      "data"),
    ("Other Income",         "other_income",                  "data"),
    ("GOSI Contribution",    "gosi_from_file",                "data"),
    ("HQ Ded",               "hq_deductions",                 "data"),
    ("Deductions",           "deductions",                    "data"),
    ("Total Deductions",     None,                            "computed:total_deductions"),
    ("Net Salary",           "net_salary_from_file",          "data"),
    ("Charge Base",          None,                            "computed:charge_base"),
    ("Hiring Date",          None,                            "emp:date_of_joining"),
    (None,                   None,                            "blank"),
    # Right-side billing block
    ("Total Salary",         "total_per_contract",            "data"),
    ("Gosi",                 None,                            "computed:gosi_billing"),
    ("ERC Fee",              None,                            "constant:erc_fee"),
    ("Bank Charges",         None,                            "constant:bank_charges"),
    ("Total",                None,                            "computed:billing_total"),
]


# Constants — until Template gets these fields wired (next chunk)
DEFAULT_ERC_FEE = 600
DEFAULT_BANK_CHARGES = 5
DEFAULT_GOSI_BILLING_RATE_SAUDI = 11.75      # employer share for Saudis
DEFAULT_GOSI_BILLING_RATE_NON_SAUDI = 2.0    # occupational hazard for non-Saudis
# GOSI is only charged on (Basic + Housing) up to this cap. Above 45K the
# contribution base is fixed at 45K.
GOSI_BILLING_CAP = 45000


def _is_saudi(emp):
    nationality = (emp or {}).get("nationality") if isinstance(emp, dict) \
        else getattr(emp, "nationality", None)
    return (nationality or "").strip().lower() in ("saudi arabia", "saudi", "ksa", "sa")


def _gosi_billing(row, emp, new_joiner=False):
    """Employer GOSI for billing: rate on (Basic+Housing), base capped at 45K.

    New joiners (first payroll) are billed GOSI on their actual worked-days
    Basic+Housing. Existing employees are billed on the FULL contract
    Basic+Housing even if they worked a partial month (so a mid-month unpaid
    leave doesn't shrink GOSI). Contract columns fall back to the worked-days
    values when not provided in the file.
    """
    rate = DEFAULT_GOSI_BILLING_RATE_SAUDI if _is_saudi(emp) \
        else DEFAULT_GOSI_BILLING_RATE_NON_SAUDI
    if new_joiner:
        base = _f(row.basic_per_working_days) + _f(row.housing_per_working_days)
    else:
        base = (_f(row.basic_per_contract) or _f(row.basic_per_working_days)) \
             + (_f(row.housing_per_contract) or _f(row.housing_per_working_days))
    return round(min(base, GOSI_BILLING_CAP) * rate / 100, 2)


def _is_new_joiner(emp_id, period_start, cache):
    """New joiner = no submitted Salary Slip ending before this payroll period."""
    if not emp_id or not period_start:
        return False
    if emp_id in cache:
        return cache[emp_id]
    has_prior = frappe.db.exists("Salary Slip", {
        "employee": emp_id, "docstatus": 1, "end_date": ["<", period_start],
    })
    cache[emp_id] = not has_prior
    return cache[emp_id]


def _bank_charges_for(row):
    """No bank transfer charge when the employee's net pay is 0 (nothing
    transferred). ERC fee + GOSI still apply."""
    return 0 if _f(row.net_salary_from_file) == 0 else DEFAULT_BANK_CHARGES

FILL_ERROR = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
FILL_WARNING = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
FILL_HEADER = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
THIN = Side(border_style="thin", color="999999")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def generate(run, template) -> str:
    """Build the internal sheet from a Payroll Import Run, attach as `internal_sheet_file`."""
    project_name = _project_display_name(template)
    month_label = _month_label(run.payroll_period_end or run.posting_date)
    wb = build_workbook(run.parsed_rows, template, project_name, month_label,
                        period_start=run.payroll_period_start)
    filename = f"RSG- {project_name} Payroll - {month_label}.xlsx".replace("  ", " ")
    return save_excel_and_link(
        wb, run.name, "internal_sheet_file", filename, "Internal Sheet",
    )


def build_workbook(rows, template, project_name, month_label, period_start=None):
    """Build the internal-sheet workbook from any iterable of row-like objects.

    Used by both the file-based Payroll Import Run and the no-file
    Payroll No File Run.  Each ``row`` only needs the same attribute surface
    that ``_resolve`` reads — see _RowLike in no_file_generator for the
    no-file shape.
    """
    rows = list(rows)
    wb = Workbook()
    ws = wb.active
    ws.title = f"RSG- {project_name.upper()}"[:31]

    ws.cell(row=1, column=1, value=f"Elite - RSP- {project_name} Payroll").font = Font(bold=True, size=12)
    ws.cell(row=2, column=1, value=f"For the month of {month_label}").font = Font(italic=True)

    for col_idx, (header, _, _) in enumerate(COLUMNS, start=1):
        if header is None:
            continue
        c = ws.cell(row=4, column=col_idx, value=header)
        c.font = Font(bold=True)
        c.fill = FILL_HEADER
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER

    from openpyxl.comments import Comment

    emp_cache = {}
    nj_cache = {}
    for serial, row in enumerate(rows, start=1):
        emp = _get_employee(getattr(row, "employee", None), emp_cache)
        excel_row = 4 + serial
        new_joiner = _is_new_joiner(getattr(row, "employee", None),
                                    period_start, nj_cache)
        iban_value, iban_warn = _iban_cell(row, emp)
        iban_col_idx = None
        for col_idx, (_, field, kind) in enumerate(COLUMNS, start=1):
            if kind == "computed:iban":
                value = iban_value
                iban_col_idx = col_idx
            else:
                value = _resolve(row, emp, template, kind, field, serial,
                                 new_joiner)
            if value is None:
                continue
            cell = ws.cell(row=excel_row, column=col_idx, value=value)
            if isinstance(value, float):
                cell.number_format = "#,##0.00"
            cell.border = BORDER

        fill = None
        status = getattr(row, "validation_status", None)
        if status == "Error":
            fill = FILL_ERROR
        elif status == "Warning":
            fill = FILL_WARNING
        if fill:
            for col_idx in range(1, len(COLUMNS) + 1):
                ws.cell(row=excel_row, column=col_idx).fill = fill
            msgs = getattr(row, "validation_messages", None)
            if msgs:
                ws.cell(row=excel_row, column=1).comment = Comment(
                    msgs[:2000], "ERC Payroll Automation"
                )

        # IBAN cell: yellow-flag a substituted/changed/invalid IBAN. Keep the
        # comment even on Error rows; only paint yellow if the row isn't red.
        if iban_warn and iban_col_idx:
            cell = ws.cell(row=excel_row, column=iban_col_idx)
            if fill is not FILL_ERROR:
                cell.fill = FILL_WARNING
            cell.comment = Comment(iban_warn[:2000], "ERC Payroll Automation")

    if rows:
        total_row_idx = 4 + len(rows) + 1
        _write_totals_row(ws, total_row_idx, first_data_row=5, last_data_row=4 + len(rows))

    for col_idx, (header, _, _) in enumerate(COLUMNS, start=1):
        if header and len(header) > 10:
            ws.column_dimensions[ws.cell(row=4, column=col_idx).column_letter].width = max(15, len(header) + 2)
        else:
            ws.column_dimensions[ws.cell(row=4, column=col_idx).column_letter].width = 14

    return wb


_TOTAL_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")


def _write_totals_row(ws, total_row_idx, first_data_row, last_data_row):
    """Write a `Total` label + SUM(...) formulas for every numeric column."""
    total_cols = {
        10, 11,           # Basic per contract, Working Days
        12, 13, 14, 15,   # Baisc, Housing, Transportaion, Other Allowance
        16, 17, 18,       # Total Salary, Overtime, Other Income
        19, 20, 21,       # GOSI, HQ Ded, Deductions
        22, 23, 24,       # Total Deductions, Net Salary, Charge Base
        27, 28, 29, 30, 31,  # Right-side billing block
    }
    label_cell = ws.cell(row=total_row_idx, column=1, value="Total")
    label_cell.font = Font(bold=True)
    label_cell.fill = _TOTAL_FILL
    label_cell.border = BORDER
    label_cell.alignment = Alignment(horizontal="right")

    for col_idx in range(1, len(COLUMNS) + 1):
        cell = ws.cell(row=total_row_idx, column=col_idx)
        cell.fill = _TOTAL_FILL
        cell.border = BORDER
        if col_idx in total_cols:
            col_letter = ws.cell(row=4, column=col_idx).column_letter
            cell.value = f"=SUM({col_letter}{first_data_row}:{col_letter}{last_data_row})"
            cell.number_format = "#,##0.00"
            cell.font = Font(bold=True)


def _resolve(row, emp, template, kind, field, serial, new_joiner=False):
    if kind == "blank":
        return None
    if kind == "data":
        return getattr(row, field, None)
    if kind.startswith("emp:"):
        fname = kind.split(":", 1)[1]
        return (emp or {}).get(fname) if isinstance(emp, dict) else getattr(emp, fname, None)
    if kind == "computed:serial":
        return serial
    if kind == "computed:company_label":
        return _project_display_name(template)
    if kind == "computed:bank_name":
        emp_bank = (emp or {}).get("bank_name") if isinstance(emp, dict) else getattr(emp, "bank_name", None)
        return bank_display_for(emp_bank, row.raw_iban) or ""
    if kind == "computed:total_deductions":
        g = _f(row.gosi_from_file); h = _f(row.hq_deductions); d = _f(row.deductions)
        return round(g + h + d, 2)
    if kind == "computed:charge_base":
        return _f(row.total_per_contract)
    if kind == "computed:gosi_billing":
        # employer-side GOSI for the billing block (base capped at 45K)
        return _gosi_billing(row, emp, new_joiner)
    if kind == "constant:erc_fee":
        return DEFAULT_ERC_FEE
    if kind == "constant:bank_charges":
        return _bank_charges_for(row)
    if kind == "computed:billing_total":
        total = _f(row.total_per_contract)
        return round(total + _gosi_billing(row, emp, new_joiner) + DEFAULT_ERC_FEE
                     + _bank_charges_for(row), 2)
    return None


def _f(v):
    try:
        return float(v) if v is not None else 0.0
    except (TypeError, ValueError):
        return 0.0


def _get_employee(emp_id, cache):
    if not emp_id:
        return None
    if emp_id in cache:
        return cache[emp_id]
    fields = ["employee_name", "employment_type", "date_of_joining",
              "nationality", "bank_name", "bank_ac_no"]
    meta = frappe.get_meta("Employee")
    for f in ("iban", "custom_iban", "bank_account_no"):
        if meta.has_field(f) and f not in fields:
            fields.append(f)
    e = frappe.db.get_value("Employee", emp_id, fields, as_dict=True)
    cache[emp_id] = e
    return e


# IBAN fields on Employee, in priority order (first non-empty wins).
_IBAN_EMP_FIELDS = ("iban", "custom_iban", "bank_ac_no", "bank_account_no")


def _clean_iban(value):
    if not value:
        return ""
    return "".join(c for c in str(value) if c.isalnum()).upper()


def _emp_iban(emp):
    if not emp:
        return ""
    for f in _IBAN_EMP_FIELDS:
        v = emp.get(f) if isinstance(emp, dict) else getattr(emp, f, None)
        if v:
            return v
    return ""


def _iban_cell(row, emp):
    """Decide the IBAN to print + an optional yellow-warning message.

    - File IBAN matches system  -> print it, no warning.
    - File IBAN invalid/missing -> use the system IBAN (if valid) + warn.
    - File IBAN valid but differs from system -> use system IBAN + warn (the
      customer may have changed it; finance should verify).
    """
    file_iban = _clean_iban(getattr(row, "raw_iban", None))
    sys_iban = _clean_iban(_emp_iban(emp))

    def valid(s):
        return bool(s) and len(s) == 24 and s.startswith("SA")

    if file_iban and sys_iban and file_iban == sys_iban:
        return file_iban, None
    if not valid(file_iban):
        if valid(sys_iban):
            return sys_iban, "File IBAN invalid/missing — used system IBAN"
        return (file_iban or sys_iban or ""), "No valid 24-char SA IBAN found"
    if sys_iban and file_iban != sys_iban:
        return sys_iban, ("IBAN in file ({0}) differs from system — used system IBAN"
                          .format(file_iban))
    return file_iban, None


def _project_display_name(template):
    """Use Project.project_name if available, else template_name."""
    if template.project:
        pn = frappe.db.get_value("Project", template.project, "project_name")
        if pn:
            return pn
    return template.template_name or "Project"


def _month_label(date_value):
    if not date_value:
        date_value = datetime.now()
    if isinstance(date_value, str):
        try:
            date_value = datetime.fromisoformat(date_value)
        except ValueError:
            return date_value
    return date_value.strftime("%B %Y")


