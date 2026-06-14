"""Payroll Posting Run engine.

Two background entry points:

  parse_sheet(run_name)  — read the approved Internal Sheet, diff fixed pay vs
                           Employee fields, build the review + variable tables.
  post(run_name)         — apply approved salary changes (Employee + new SSA),
                           create submitted Additional Salary, create a DRAFT
                           Payroll Entry.

The Internal Sheet is OUR own 33-column RSG layout (see generators/internal_sheet
COLUMNS), so parsing is deterministic — we match by the row-4 header text.
"""

import json

import frappe
from frappe.utils import getdate, now_datetime, flt

from ..parser.file_parser import _resolve_attached_file_path


# Internal-sheet header text  ->  logical key
HEADER_MAP = {
    "erp id no.": "employee",
    "basic per contract": "contract_basic",
    "working days": "working_days",
    "baisc": "wd_basic",
    "housing": "wd_housing",
    "transportaion": "wd_transport",
    "other allowance": "wd_other",
    "overtime": "overtime",
    "other income": "other_income",
    "hq ded": "hq_deductions",
    "deductions": "deductions",
}

# Employee fixed-pay field  ->  (sheet key, human label, prorated?)
FIXED_PAY = [
    ("basic_salary",       "contract_basic", "Basic Salary",            False),
    ("housing_allowance",  "wd_housing",     "Housing Allowance",       True),
    ("transport_allowance", "wd_transport",  "Transport Allowance",     True),
    ("food_allowance",     "wd_other",       "Food / Other Allowance",  True),
]

# Sheet variable key  ->  ERPNext Salary Component (verified against site)
VARIABLE_COMPONENTS = [
    ("overtime",      "Overtime"),
    ("other_income",  "Other Additions"),
    ("deductions",    "Deductions"),
    ("hq_deductions", "HQ Deductions"),
]


# ---------------------------------------------------------------------------
# Parse
# ---------------------------------------------------------------------------

def parse_sheet(run_name, user=None):
    frappe.log_error(
        title=f"Payroll Posting parse STARTED: {run_name}",
        message=f"user={user}",
    )
    _set_user(user)
    run = frappe.get_doc("Payroll Posting Run", run_name)
    try:
        rows = _read_sheet(run)
        period_days = _period_days(run)

        salary_changes = []
        variable_preview = []
        emp_cache = {}

        for r in rows:
            emp_id = r.get("employee")
            if not emp_id:
                continue
            emp = _get_employee(emp_id, emp_cache)
            if not emp:
                # Unknown employee in the approved sheet — surface, don't crash
                frappe.log_error(
                    title=f"Payroll Posting: unknown employee {run_name}",
                    message=f"Sheet row references {emp_id!r} which is not an Employee. Skipped.",
                )
                continue

            working_days = flt(r.get("working_days")) or 0
            prorated = bool(working_days and working_days < period_days)

            for fieldname, sheet_key, label, can_prorate in FIXED_PAY:
                sheet_val = r.get(sheet_key)
                if sheet_val is None or sheet_val == "":
                    continue
                sheet_val = flt(sheet_val)
                cur_val = flt(emp.get(fieldname))
                if abs(cur_val - sheet_val) <= 0.01:
                    continue
                is_prorated = prorated and can_prorate
                salary_changes.append({
                    "employee": emp_id,
                    "employee_name": emp.get("employee_name"),
                    "field_label": label,
                    "fieldname": fieldname,
                    "current_value": cur_val,
                    "sheet_value": sheet_val,
                    "prorated_flag": 1 if is_prorated else 0,
                    # Auto-tick contract Basic always; allowances only when not prorated
                    "apply": 0 if is_prorated else 1,
                })

            for sheet_key, component in VARIABLE_COMPONENTS:
                amt = flt(r.get(sheet_key))
                if abs(amt) < 0.01:
                    continue
                variable_preview.append({
                    "employee": emp_id,
                    "employee_name": emp.get("employee_name"),
                    "salary_component": component,
                    "amount": amt,
                })

        run.set("salary_changes", salary_changes)
        run.set("additional_salary_preview", variable_preview)
        run.employees_in_sheet = len({r["employee"] for r in rows if r.get("employee")})
        run.salary_changes_count = len(salary_changes)
        run.additional_salary_count = len(variable_preview)
        run.status = "Review"
        run.posting_log = (
            f"Parsed {run.employees_in_sheet} employees. "
            f"{len(salary_changes)} fixed-pay changes detected "
            f"({sum(1 for c in salary_changes if c['prorated_flag'])} prorated — left unticked). "
            f"{len(variable_preview)} Additional Salary rows queued."
        )
        run.save(ignore_permissions=True)
        frappe.db.commit()

        frappe.log_error(
            title=f"Payroll Posting parse COMPLETE: {run_name}",
            message=run.posting_log,
        )
    except Exception:
        frappe.log_error(
            title=f"Payroll Posting parse FAILED: {run_name}",
            message=frappe.get_traceback(),
        )
        run.db_set("status", "Failed", update_modified=False)
        run.db_set("posting_log", frappe.get_traceback())
        frappe.db.commit()
        raise


# ---------------------------------------------------------------------------
# Post
# ---------------------------------------------------------------------------

def post(run_name, user=None):
    frappe.log_error(
        title=f"Payroll Posting post STARTED: {run_name}",
        message=f"user={user}",
    )
    _set_user(user)
    run = frappe.get_doc("Payroll Posting Run", run_name)
    log_lines = []
    try:
        # Resolve the Salary Structure from the Template (source of truth).
        # run.salary_structure is fetched at save-time and can be stale if the
        # Template was filled in AFTER this run was created — so fall back to
        # the live template value and persist it back onto the run.
        if not run.salary_structure and run.template:
            tmpl_ss = frappe.db.get_value(
                "Payroll Import Template", run.template, "salary_structure")
            if tmpl_ss:
                run.salary_structure = tmpl_ss
                run.db_set("salary_structure", tmpl_ss, update_modified=False)
        if not run.salary_structure:
            raise ValueError(
                "Template '{0}' has no Salary Structure set. Open the Template "
                "and fill in the 'Salary Structure' field before posting.".format(
                    run.template)
            )

        period_start = getdate(run.payroll_period_start)
        # Additional Salary must fall INSIDE the payroll month, otherwise a
        # record can leak into the wrong month's slip. Clamp posting_date into
        # [period_start, period_end]; default to period_end.
        period_end = getdate(run.payroll_period_end)
        payroll_date = getdate(run.posting_date) if run.posting_date else period_end
        if payroll_date < period_start or payroll_date > period_end:
            payroll_date = period_end

        # ---- 1. Apply approved fixed-pay changes (Employee + new SSA) --------
        # Group approved changes by employee
        by_emp = {}
        for ch in run.salary_changes:
            if not ch.apply:
                continue
            by_emp.setdefault(ch.employee, {})[ch.fieldname] = flt(ch.sheet_value)

        ssa_created = ssa_skipped = emp_updated = 0
        for emp_id, changes in by_emp.items():
            try:
                # Update Employee fixed-pay fields directly
                frappe.db.set_value("Employee", emp_id, changes,
                                    update_modified=True)
                emp_updated += 1

                # A base change (or no active SSA) requires a fresh dated SSA
                base_changed = "basic_salary" in changes
                made = _ensure_ssa(
                    emp_id, run.salary_structure, period_start,
                    new_base=changes.get("basic_salary"),
                    allowances={
                        k: v for k, v in changes.items()
                        if k in ("housing_allowance", "transport_allowance",
                                 "food_allowance")
                    },
                    project=run.project,
                    force=base_changed,
                    log_lines=log_lines,
                )
                if made == "created":
                    ssa_created += 1
                elif made == "skipped":
                    ssa_skipped += 1
            except Exception as e:
                log_lines.append(f"[SSA ERROR] {emp_id}: {e}")
                frappe.log_error(
                    title=f"Payroll Posting: SSA error {emp_id} ({run_name})",
                    message=frappe.get_traceback(),
                )

        frappe.db.commit()

        # ---- 2. Submitted Additional Salary for variable components ---------
        addsal_created = 0
        for row in run.additional_salary_preview:
            try:
                _create_additional_salary(
                    row.employee, row.salary_component, flt(row.amount),
                    payroll_date, period_start, run.payroll_period_end,
                )
                addsal_created += 1
            except Exception as e:
                log_lines.append(
                    f"[ADDSAL ERROR] {row.employee} {row.salary_component}: {e}"
                )
                frappe.log_error(
                    title=f"Payroll Posting: Additional Salary error ({run_name})",
                    message=frappe.get_traceback(),
                )
        frappe.db.commit()

        # ---- 3. Draft Payroll Entry ----------------------------------------
        pe_name = None
        try:
            pe_name = _create_draft_payroll_entry(run)
            if pe_name:
                run.db_set("created_payroll_entry", pe_name, update_modified=False)
        except Exception as e:
            log_lines.append(f"[PAYROLL ENTRY ERROR] {e}")
            frappe.log_error(
                title=f"Payroll Posting: Payroll Entry error ({run_name})",
                message=frappe.get_traceback(),
            )

        summary = (
            f"Employees updated: {emp_updated}. "
            f"SSAs created: {ssa_created}, skipped (unchanged): {ssa_skipped}. "
            f"Additional Salary submitted: {addsal_created}. "
            f"Draft Payroll Entry: {pe_name or 'NOT created — see log'}."
        )
        run.db_set("status", "Posted", update_modified=False)
        run.db_set("posting_log", summary + "\n\n" + "\n".join(log_lines),
                   update_modified=False)
        frappe.db.commit()

        frappe.log_error(
            title=f"Payroll Posting post COMPLETE: {run_name}",
            message=summary,
        )
    except Exception:
        frappe.log_error(
            title=f"Payroll Posting post FAILED: {run_name}",
            message=frappe.get_traceback(),
        )
        run.db_set("status", "Failed", update_modified=False)
        run.db_set(
            "posting_log",
            (frappe.get_traceback() + "\n\n" + "\n".join(log_lines)),
            update_modified=False,
        )
        frappe.db.commit()
        raise


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _set_user(user):
    eff = user or frappe.session.user or "Administrator"
    if eff in ("Guest", "", None):
        eff = "Administrator"
    frappe.set_user(eff)


def _period_days(run):
    try:
        d = (getdate(run.payroll_period_end) - getdate(run.payroll_period_start)).days + 1
        return d if d > 0 else 30
    except Exception:
        return 30


def _read_sheet(run):
    """Read the approved Internal Sheet into a list of dicts keyed by HEADER_MAP."""
    from openpyxl import load_workbook

    path = _resolve_attached_file_path(run.approved_sheet)
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]

    all_rows = list(ws.iter_rows(values_only=True))
    wb.close()

    # Find the header row (the one containing "ERP ID No.")
    header_idx = None
    for i, row in enumerate(all_rows[:10]):
        norm = [str(c).strip().lower() if c is not None else "" for c in row]
        if "erp id no." in norm:
            header_idx = i
            break
    if header_idx is None:
        raise ValueError(
            "Could not find the 'ERP ID No.' header in the first 10 rows — "
            "is this an Internal Sheet produced by a Payroll Import Run?"
        )

    headers = [str(c).strip().lower() if c is not None else ""
               for c in all_rows[header_idx]]
    col_of = {}
    for col_idx, h in enumerate(headers):
        key = HEADER_MAP.get(h)
        if key and key not in col_of:   # first occurrence wins (left block)
            col_of[key] = col_idx

    if "employee" not in col_of:
        raise ValueError("Internal Sheet has no 'ERP ID No.' column.")

    out = []
    for row in all_rows[header_idx + 1:]:
        emp = row[col_of["employee"]] if col_of["employee"] < len(row) else None
        if emp is None or str(emp).strip() == "":
            continue
        if str(emp).strip().lower() == "total":   # totals row
            continue
        rec = {}
        for key, ci in col_of.items():
            rec[key] = row[ci] if ci < len(row) else None
        rec["employee"] = str(rec["employee"]).strip()
        out.append(rec)
    return out


def _get_employee(emp_id, cache):
    if emp_id in cache:
        return cache[emp_id]
    e = frappe.db.get_value(
        "Employee", emp_id,
        ["name", "employee_name", "department", "designation", "company",
         "basic_salary", "housing_allowance", "transport_allowance",
         "food_allowance"],
        as_dict=True,
    )
    cache[emp_id] = e
    return e


def _ensure_ssa(emp_id, salary_structure, from_date, new_base, allowances,
                project, force, log_lines):
    """Create a dated Salary Structure Assignment when base changed or none
    exists. Idempotent for re-runs of the same period.

    Returns "created" | "skipped".
    """
    # Latest submitted SSA for this employee (any date)
    latest = frappe.db.get_value(
        "Salary Structure Assignment",
        {"employee": emp_id, "docstatus": 1},
        ["name", "base", "from_date", "company", "currency",
         "payroll_payable_account"],
        order_by="from_date desc",
        as_dict=True,
    )

    target_base = flt(new_base) if new_base is not None else (
        flt(latest.base) if latest else 0
    )

    # Nothing to do: an SSA exists, base unchanged, and not forced
    if latest and not force and abs(flt(latest.base) - target_base) <= 0.01:
        return "skipped"

    # Same-period SSA already there? There can be MORE THAN ONE submitted SSA
    # with the same from_date (messy historical data), and HRMS throws
    # DuplicateAssignment on insert if even one survives — so cancel them ALL.
    existing_same_date = frappe.get_all(
        "Salary Structure Assignment",
        filters={"employee": emp_id, "from_date": from_date, "docstatus": 1},
        fields=["name", "base"],
    )
    if (len(existing_same_date) == 1
            and abs(flt(existing_same_date[0].base) - target_base) <= 0.01):
        return "skipped"
    for ex in existing_same_date:
        old = frappe.get_doc("Salary Structure Assignment", ex.name)
        old.cancel()
        log_lines.append(
            f"[SSA] {emp_id}: cancelled {old.name} "
            f"(base {old.base} -> {target_base}) for re-post."
        )

    emp = frappe.db.get_value(
        "Employee", emp_id,
        ["employee_name", "department", "designation", "company"],
        as_dict=True,
    )
    company = (latest.company if latest else None) or emp.company \
        or frappe.defaults.get_global_default("company")
    currency = (latest.currency if latest else None) \
        or frappe.db.get_value("Company", company, "default_currency") or "SAR"
    ppa = (latest.payroll_payable_account if latest else None) \
        or frappe.db.get_value("Company", company,
                               "default_payroll_payable_account")

    ssa = frappe.new_doc("Salary Structure Assignment")
    ssa.employee = emp_id
    ssa.employee_name = emp.employee_name
    ssa.department = emp.department
    ssa.designation = emp.designation
    ssa.salary_structure = salary_structure
    ssa.from_date = from_date
    ssa.company = company
    ssa.currency = currency
    if ppa:
        ssa.payroll_payable_account = ppa
    ssa.base = target_base
    ssa.variable = 0
    # Mirror allowances onto the SSA so the structure formula resolves them
    # regardless of Employee-vs-SSA precedence on this site.
    for k, v in (allowances or {}).items():
        if ssa.meta.has_field(k):
            ssa.set(k, flt(v))
    if project and ssa.meta.has_field("project"):
        ssa.project = project
    ssa.flags.ignore_permissions = True
    ssa.insert(ignore_permissions=True)
    ssa.submit()
    log_lines.append(
        f"[SSA] {emp_id}: created {ssa.name} structure={salary_structure} "
        f"base={target_base} from={from_date}"
    )
    return "created"


def _create_additional_salary(emp_id, component, amount, payroll_date,
                              period_start=None, period_end=None):
    company = frappe.db.get_value("Employee", emp_id, "company") \
        or frappe.defaults.get_global_default("company")

    # Supersede any existing 'overwrite' Additional Salary for the same
    # employee + component in this payroll period. Without this, re-posting a
    # period (or posting it from more than one run) leaves multiple overwrite
    # rows, and HRMS then refuses slip creation:
    #   "Multiple Additional Salaries with overwrite property exist for
    #    Salary Component X between <start> and <end>".
    if period_start and period_end:
        for name in frappe.get_all(
                "Additional Salary",
                filters={
                    "employee": emp_id,
                    "salary_component": component,
                    "overwrite_salary_structure_amount": 1,
                    "docstatus": 1,
                    "payroll_date": ["between", [period_start, period_end]],
                },
                pluck="name"):
            frappe.get_doc("Additional Salary", name).cancel()

    doc = frappe.new_doc("Additional Salary")
    doc.employee = emp_id
    doc.salary_component = component
    doc.amount = abs(amount)
    doc.payroll_date = payroll_date
    doc.company = company
    doc.overwrite_salary_structure_amount = 1
    if doc.meta.has_field("currency"):
        doc.currency = frappe.db.get_value("Company", company,
                                           "default_currency") or "SAR"
    doc.flags.ignore_permissions = True
    doc.insert(ignore_permissions=True)
    doc.submit()
    return doc.name


def _create_draft_payroll_entry(run):
    """Create a DRAFT Payroll Entry scoped to the run's project + period, with
    employees already populated, left in workflow state 'Draft'.

    Finance then runs Actions -> Create Salary Slips and the HRM/FM approval
    chain manually (the site's Payroll Entry workflow). We do NOT create or
    submit Salary Slips here.

    Employees are scoped HERE (project + work location + active SSA) — we do
    NOT use pe.get_emp_list(), because that depends on a manual HRMS core patch
    (make_filters/get_filter_condition adding `projects`) that Frappe Cloud
    wipes on every rebuild; when missing it returns every active employee
    company-wide.

    Raises on a missing payable account or no eligible employees, so the
    failure shows on the Posting Run instead of leaving a silently-empty PE.
    """
    # Prefer the company used by employees in this run; fall back to global default
    company = frappe.defaults.get_global_default("company")
    if run.salary_changes:
        c = frappe.db.get_value("Employee", run.salary_changes[0].employee,
                                "company")
        if c:
            company = c
    elif run.additional_salary_preview:
        c = frappe.db.get_value("Employee",
                                run.additional_salary_preview[0].employee,
                                "company")
        if c:
            company = c
    if not company:
        raise ValueError("Could not determine the Company for the Payroll Entry.")

    # Payroll Payable Account is mandatory on Payroll Entry — resolve from the
    # Company default and fail loudly if it isn't set.
    payable_account = frappe.db.get_value(
        "Company", company, "default_payroll_payable_account")
    if not payable_account:
        raise ValueError(
            f"Company '{company}' has no Default Payroll Payable Account. "
            f"Set it under Company > Accounting Settings before posting."
        )
    currency = frappe.db.get_value("Company", company, "default_currency") or "SAR"

    # Work-location scope comes from the template (the run isn't location-aware).
    work_location, location_field = None, "custom_location"
    if run.template:
        t = frappe.db.get_value(
            "Payroll Import Template", run.template,
            ["filter_by_work_location", "location_field_on_employee"],
            as_dict=True) or {}
        work_location = t.get("filter_by_work_location")
        location_field = t.get("location_field_on_employee") or "custom_location"

    pe = frappe.new_doc("Payroll Entry")
    pe.company = company
    pe.posting_date = run.posting_date or run.payroll_period_end
    pe.payroll_frequency = "Monthly"
    pe.start_date = run.payroll_period_start
    pe.end_date = run.payroll_period_end
    pe.payroll_payable_account = payable_account
    if pe.meta.has_field("currency"):
        pe.currency = currency
    # SAR == company currency, so the exchange rate is 1. Mandatory on PE.
    if pe.meta.has_field("exchange_rate"):
        pe.exchange_rate = 1
    # The live site stores the client/project scope on `projects` (the custom
    # field the patched core filters on). `project` is a leftover standard
    # field — only used as a fallback if `projects` is absent.
    if run.project and pe.meta.has_field("projects"):
        pe.projects = run.project
    elif run.project and pe.meta.has_field("project"):
        pe.project = run.project
    # Land in the workflow's initial state so finance's normal buttons appear.
    if pe.meta.has_field("workflow_state"):
        pe.workflow_state = "Draft"
    pe.flags.ignore_permissions = True
    pe.insert(ignore_permissions=True)

    # The Payroll Entry must follow the approved file: scope to exactly the
    # employees in the sheet. Fall back to project scope only if the sheet
    # can't be read.
    sheet_emp_ids = _sheet_employee_ids(run) or None

    # Ensure every sheet employee has a Salary Structure Assignment whose base
    # equals the sheet's WORKED-DAYS Basic — the amount the slip's Basic must
    # show. The salary structure's Basic component is `base` and is NOT
    # payment-day prorated, so a bridging joiner's extra days reach the slip
    # ONLY through the SSA base; sourcing it from the contract Basic leaves
    # bridges short. Also creates SSAs for new hires that have none.
    emp_wd_basic = _sheet_worked_days_basic(run)
    _ensure_payroll_ssas(run, company, work_location, location_field,
                         emp_ids=sheet_emp_ids, emp_wd_basic=emp_wd_basic)

    employees = _eligible_employees(run, company, work_location, location_field,
                                    emp_ids=sheet_emp_ids)
    if not employees:
        # Don't leave an empty shell PE behind.
        frappe.delete_doc("Payroll Entry", pe.name,
                          ignore_permissions=True, force=True)
        scope = ("the approved sheet" if sheet_emp_ids
                 else f"project '{run.project}'")
        raise ValueError(
            f"No eligible employees found from {scope} in "
            f"{run.payroll_period_start} - {run.payroll_period_end}. "
            f"Check that the employees have an active Salary Structure "
            f"Assignment covering this period."
        )
    for e in employees:
        pe.append("employees", {
            "employee": e.name,
            "employee_name": e.employee_name,
            "department": e.department,
            "designation": e.designation,
        })
    pe.number_of_employees = len(pe.employees)
    pe.save(ignore_permissions=True)
    return pe.name


def _sheet_employee_ids(run):
    """Employee IDs present in the approved Internal Sheet, so the Payroll
    Entry follows the file. Returns [] if the sheet can't be read (caller
    falls back to project scope)."""
    if not run.approved_sheet:
        return []
    try:
        rows = _read_sheet(run)
    except Exception:
        frappe.log_error(
            title=f"Payroll Posting: sheet re-read for PE scope failed ({run.name})",
            message=frappe.get_traceback())
        return []
    out, seen = [], set()
    for r in rows:
        eid = str(r.get("employee") or "").strip()
        if eid and eid not in seen and frappe.db.exists("Employee", eid):
            seen.add(eid)
            out.append(eid)
    return out


def _sheet_worked_days_basic(run):
    """Map each sheet employee -> the WORKED-DAYS Basic (the sheet's "Baisc"
    column = contract Basic x working_days / period_days). This is the amount
    the slip's Basic must equal, so it becomes the SSA base. Falls back to the
    contract Basic when the worked-days cell is blank. Empty dict if the sheet
    can't be read (callers then fall back to Employee.basic_salary)."""
    if not run.approved_sheet:
        return {}
    try:
        rows = _read_sheet(run)
    except Exception:
        frappe.log_error(
            title=f"Payroll Posting: sheet re-read for SSA base failed ({run.name})",
            message=frappe.get_traceback())
        return {}
    out = {}
    for r in rows:
        eid = str(r.get("employee") or "").strip()
        if not eid:
            continue
        wd_basic = flt(r.get("wd_basic"))
        contract_basic = flt(r.get("contract_basic"))
        val = wd_basic if wd_basic > 0 else contract_basic
        if val > 0:
            out[eid] = val
    return out


def _ensure_payroll_ssas(run, company, work_location=None,
                         location_field="custom_location", emp_ids=None,
                         emp_wd_basic=None):
    """Ensure every in-scope employee has a submitted Salary Structure
    Assignment whose ``base`` equals the sheet's WORKED-DAYS Basic — the amount
    the slip's Basic component (formula ``base``, not payment-day prorated) must
    show. Creates a dated SSA when one is missing OR when the existing base
    differs (e.g. a bridging joiner whose worked-days Basic exceeds the
    contract Basic); skips employees already on the right base, so normal
    (working_days == period) employees are untouched. Returns the count
    created/updated.

    When ``emp_ids`` is given (the sheet's employees) those are used; otherwise
    falls back to the project/work-location scope. ``emp_wd_basic`` maps
    employee -> worked-days Basic (from the sheet); employees without an entry
    fall back to Employee.basic_salary. No-op if the run has no Salary Structure
    (trigger_post already blocks posting in that case).
    """
    salary_structure = run.salary_structure
    if not salary_structure:
        return 0
    emp_wd_basic = emp_wd_basic or {}

    if emp_ids is not None:
        emps = list(emp_ids)
    else:
        emp_meta = frappe.get_meta("Employee")
        filters = {"status": "Active"}
        if company and emp_meta.has_field("company"):
            filters["company"] = company
        if run.project and emp_meta.has_field("project"):
            filters["project"] = run.project
        if work_location and emp_meta.has_field(location_field):
            filters[location_field] = work_location
        emps = frappe.get_all("Employee", filters=filters, pluck="name")
    if not emps:
        return 0

    from_date = getdate(run.payroll_period_start)
    created = 0
    for emp_id in emps:
        # The slip's Basic must equal the worked-days Basic from the sheet; fall
        # back to the Employee's standing basic pay when the sheet has no value.
        target_base = flt(emp_wd_basic.get(emp_id))
        if target_base <= 0:
            target_base = flt(frappe.db.get_value("Employee", emp_id, "basic_salary"))
        if target_base <= 0:
            frappe.log_error(
                title=f"Payroll Posting: no base for SSA {emp_id} ({run.name})",
                message=f"{emp_id} has no worked-days/basic salary — SSA not set.",
            )
            continue
        try:
            # force=False -> _ensure_ssa skips when the latest base already
            # matches (no churn for unchanged employees) and creates a dated SSA
            # only when missing or the base differs.
            made = _ensure_ssa(emp_id, salary_structure, from_date,
                               new_base=target_base, allowances={},
                               project=run.project, force=False, log_lines=[])
            if made == "created":
                created += 1
        except Exception:
            frappe.log_error(
                title=f"Payroll Posting: SSA base sync failed {emp_id} ({run.name})",
                message=frappe.get_traceback(),
            )

    if created:
        frappe.db.commit()
        frappe.log_error(
            title=f"Payroll Posting: synced {created} SSA base(s) ({run.name})",
            message=f"Created/updated {created} Salary Structure Assignment "
                    f"base(s) to the sheet's worked-days Basic.",
        )
    return created


def _eligible_employees(run, company, work_location=None,
                        location_field="custom_location", emp_ids=None):
    """Employees to put on the Payroll Entry that have a submitted Salary
    Structure Assignment covering the period.

    When ``emp_ids`` is given (the approved sheet's employees) the PE follows
    the file exactly; otherwise falls back to the project/work-location scope.
    Built explicitly (not via pe.get_emp_list) so scoping does NOT depend on the
    manual HRMS core patch that Frappe Cloud wipes on rebuild.
    """
    if emp_ids is not None:
        if not emp_ids:
            return []
        emps = frappe.get_all(
            "Employee", filters={"name": ["in", list(emp_ids)]},
            fields=["name", "employee_name", "department", "designation"],
            order_by="employee_name asc",
        )
    else:
        emp_meta = frappe.get_meta("Employee")
        filters = {"status": "Active"}
        if company and emp_meta.has_field("company"):
            filters["company"] = company
        if run.project and emp_meta.has_field("project"):
            filters["project"] = run.project
        if work_location and emp_meta.has_field(location_field):
            filters[location_field] = work_location
        emps = frappe.get_all(
            "Employee", filters=filters,
            fields=["name", "employee_name", "department", "designation"],
            order_by="employee_name asc",
        )
    if not emps:
        return []

    names = [e.name for e in emps]
    ssa_emps = {
        r.employee for r in frappe.get_all(
            "Salary Structure Assignment",
            filters={"employee": ["in", names], "docstatus": 1,
                     "from_date": ["<=", run.payroll_period_end]},
            fields=["employee"], distinct=True,
        )
    }
    return [e for e in emps if e.name in ssa_emps]
