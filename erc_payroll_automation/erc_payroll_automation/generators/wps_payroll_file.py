"""Generate the Al Rajhi WPS payroll upload workbook from ERPNext Salary Slips."""

import io
import os
import re
from collections import defaultdict

import frappe
from frappe import _
from frappe.utils import flt, getdate, now_datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

from ..parser.bank_codes import bank_display_for


TEMPLATE_FILENAME = "psh_wps_payroll_upload_template.xlsm"
SHEET_NAME = "Sheet1"
DATA_START_ROW = 8
MAX_TEMPLATE_DATA_ROW = 65243
VALIDATION_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
IQAMA_VALIDATION_FILL = PatternFill(start_color="D9EAF7", end_color="D9EAF7", fill_type="solid")

BASIC_COMPONENT_ALIASES = ("Basic Salary", "Basic")
HOUSING_COMPONENT_ALIASES = (
    "Housing Allowance",
    "Housing",
    "House Rent Allowance",
    "HRA",
)    


def generate_wps_file(run_name: str, user: str = None):
    frappe.log_error(
        title=f"Payroll WPS generate STARTED: {run_name}",
        message=f"Worker picked up WPS generate job (user={user})",
    )

    effective_user = user or frappe.session.user or "Administrator"
    if effective_user in ("Guest", "", None):
        effective_user = "Administrator"
    frappe.set_user(effective_user)

    run = frappe.get_doc("Payroll WPS Export Run", run_name)

    try:
        slips = _get_salary_slips(run)
        if not slips:
            frappe.throw(_(
                "No submitted Salary Slips found for the selected Project, "
                "Department, Work Location, and payroll period."
            ))

        rows, warning_lines = _build_wps_rows(run, slips)
        if not rows:
            frappe.throw(_("No employees could be prepared for the WPS file."))
        if len(rows) > (MAX_TEMPLATE_DATA_ROW - DATA_START_ROW + 1):
            frappe.throw(_("Too many employees for the WPS template: {0}").format(len(rows)))

        wb = _build_workbook(rows)
        project_label = _project_label(run.project)
        department_label = _department_label(run.department)
        work_location_label = _work_location_label(getattr(run, "work_location", None))
        period_label = _period_label(run.payroll_period_start, run.payroll_period_end)
        filename = f"WPS Payroll - {project_label}"
        if department_label:
            filename += f" - {department_label}"
        if work_location_label:
            filename += f" - {work_location_label}"
        filename += f" - {period_label}.xlsx"

        file_url = _save_workbook_and_link(
            wb=wb,
            doctype="Payroll WPS Export Run",
            docname=run.name,
            fieldname="wps_file",
            filename=filename,
            label="WPS Payroll Upload File",
        )

        log_lines = [
            f"Generated {len(rows)} employee row(s) from {len(slips)} submitted Salary Slip(s).",
            f"Project: {project_label}",
        ]
        if department_label:
            log_lines.append(f"Department: {department_label}")
        if work_location_label:
            log_lines.append(f"Work Location: {work_location_label}")
        if warning_lines:
            log_lines.append("")
            log_lines.append("Warnings:")
            log_lines.extend(warning_lines)

        frappe.db.set_value(
            "Payroll WPS Export Run",
            run.name,
            {
                "status": "Generated",
                "salary_slips_included": len(slips),
                "employees_included": len(rows),
                "generated_at": now_datetime(),
                "generated_by": effective_user,
                "generation_log": "\n".join(log_lines),
            },
            update_modified=False,
        )
        frappe.db.commit()

        frappe.log_error(
            title=f"Payroll WPS generate COMPLETE: {run_name}",
            message=f"employees={len(rows)} slips={len(slips)} file_url={file_url!r}",
        )

    except Exception:
        log_message = frappe.get_traceback()
        frappe.log_error(
            title=f"Payroll WPS generate FAILED: {run_name}",
            message=frappe.get_traceback(),
        )
        try:
            frappe.db.set_value(
                "Payroll WPS Export Run",
                run.name,
                {
                    "status": "Failed",
                    "generation_log": log_message,
                },
                update_modified=False,
            )
            frappe.db.commit()
        except Exception:
            pass
        raise


def _get_salary_slips(run):
    _require_field("Salary Slip", "start_date")
    _require_field("Salary Slip", "end_date")

    slip_meta = frappe.get_meta("Salary Slip")
    employee_meta = frappe.get_meta("Employee")
    base_filters = [
        ["docstatus", "=", 1],
        ["start_date", ">=", getdate(run.payroll_period_start)],
        ["end_date", "<=", getdate(run.payroll_period_end)],
    ]
    if run.department and slip_meta.has_field("department"):
        base_filters.append(["department", "=", run.department])

    employee_names = _employees_for_scope(run, employee_meta)
    employee_filter_needed = getattr(run, "work_location", None) or not slip_meta.has_field("project")
    if employee_filter_needed and not employee_names:
        return []

    fields = _safe_fields(
        "Salary Slip",
        [
            "name",
            "employee",
            "employee_name",
            "department",
            "start_date",
            "end_date",
            "gross_pay",
            "total_deduction",
            "net_pay",
            "bank_name",
            "bank_account_no",
            "bank_ac_no",
            "employee_no",
            "employee_number",
            "project",
        ],
    )

    slips_by_name = {}

    if slip_meta.has_field("project"):
        filters = base_filters + [["project", "=", run.project]]
        if getattr(run, "work_location", None):
            for chunk in _chunks(employee_names, 500):
                for slip in frappe.get_all(
                    "Salary Slip",
                    filters=filters + [["employee", "in", chunk]],
                    fields=fields,
                    order_by="employee_name asc, employee asc",
                ):
                    slips_by_name[slip.name] = slip
        else:
            for slip in frappe.get_all(
                "Salary Slip",
                filters=filters,
                fields=fields,
                order_by="employee_name asc, employee asc",
            ):
                slips_by_name[slip.name] = slip
    elif employee_meta.has_field("project"):
        for chunk in _chunks(employee_names, 500):
            for slip in frappe.get_all(
                "Salary Slip",
                filters=base_filters + [["employee", "in", chunk]],
                fields=fields,
                order_by="employee_name asc, employee asc",
            ):
                slips_by_name[slip.name] = slip

    if not slip_meta.has_field("project") and not employee_meta.has_field("project"):
        frappe.throw(_(
            "Cannot filter by Project because neither Salary Slip.project nor "
            "Employee.project exists on this site."
        ))

    return sorted(
        slips_by_name.values(),
        key=lambda d: ((d.get("employee_name") or ""), (d.get("employee") or ""), d.get("name")),
    )


def _employees_for_scope(run, employee_meta):
    filters = {}
    if employee_meta.has_field("project"):
        filters["project"] = run.project
    if run.department and employee_meta.has_field("department"):
        filters["department"] = run.department
    if getattr(run, "work_location", None):
        if not employee_meta.has_field("custom_location"):
            frappe.throw(_("Employee.custom_location field is required to filter by Work Location"))
        filters["custom_location"] = run.work_location

    if not filters:
        return []
    return frappe.get_all("Employee", filters=filters, pluck="name")


def _build_wps_rows(run, slips):
    slip_names = [s.name for s in slips]
    detail_totals = _get_salary_detail_totals(slip_names)
    employees = _get_employee_map([s.employee for s in slips if s.get("employee")])

    basic_keys = _component_keys(run.basic_salary_component, BASIC_COMPONENT_ALIASES)
    housing_keys = _component_keys(run.housing_allowance_component, HOUSING_COMPONENT_ALIASES)

    by_employee = {}
    warning_lines = []
    for slip in slips:
        employee_id = slip.get("employee")
        if _is_zero_net_pay(slip):
            warning_lines.append(
                "{0}: skipped because Net Pay is 0".format(
                    slip.get("employee_name") or employee_id
                )
            )
            continue

        emp = employees.get(employee_id) or {}
        detail = detail_totals.get(slip.name, {})
        earnings = detail.get("earnings", {})
        deductions_by_component = detail.get("deductions", {})

        basic = _pick_component_total(earnings, basic_keys)
        housing = _pick_component_total(earnings, housing_keys)
        basic_found = _has_component(earnings, basic_keys)
        housing_found = _has_component(earnings, housing_keys)
        gross_pay = flt(slip.get("gross_pay")) or sum(earnings.values())
        deductions = _deduction_amount(slip, gross_pay, deductions_by_component)
        other_earnings = round(gross_pay - basic - housing, 2)
        raw_employee_name = slip.get("employee_name") or emp.get("employee_name") or ""
        clean_employee_name = _clean_employee_name(raw_employee_name)
        iban = _iban(slip, emp)
        iban_error = _iban_error(iban)
        national_id = _national_id(emp)
        iqama_error = _iqama_error(national_id)
        if gross_pay and not basic_found:
            warning_lines.append(
                "{0}: Basic component was not detected. Earnings found: {1}".format(
                    raw_employee_name or employee_id,
                    _component_names_for_log(earnings),
                )
            )
        if gross_pay and not housing_found:
            warning_lines.append(
                "{0}: Housing component was not detected. Earnings found: {1}".format(
                    raw_employee_name or employee_id,
                    _component_names_for_log(earnings),
                )
            )

        rec = by_employee.setdefault(
            employee_id,
            {
                "bank_name": _bank_name(slip, emp),
                "iban": iban,
                "employee_name": clean_employee_name,
                "raw_employee_name": raw_employee_name,
                "employee_number": _employee_number(slip, emp, employee_id),
                "national_id": national_id,
                "has_iqama_issue": False,
                "validation_issues": [],
                "basic": 0.0,
                "housing": 0.0,
                "other_earnings": 0.0,
                "deductions": 0.0,
                "department": _department_value(slip, emp),
                "salary_slips": [],
            },
        )
        if iban_error and iban_error not in rec["validation_issues"]:
            rec["validation_issues"].append(f"IBAN issue: {iban_error}")
        if iqama_error and iqama_error not in rec["validation_issues"]:
            rec["validation_issues"].append(f"Iqama issue: {iqama_error}")
            rec["has_iqama_issue"] = True
        rec["basic"] = round(rec["basic"] + basic, 2)
        rec["housing"] = round(rec["housing"] + housing, 2)
        rec["other_earnings"] = round(rec["other_earnings"] + other_earnings, 2)
        rec["deductions"] = round(rec["deductions"] + deductions, 2)
        rec["salary_slips"].append(slip.name)

    rows = []
    for employee_id, rec in sorted(
        by_employee.items(),
        key=lambda item: ((item[1].get("employee_name") or ""), item[0] or ""),
    ):
        missing = []
        if not rec["bank_name"]:
            missing.append("bank name")
        if not rec["national_id"]:
            missing.append("national ID")
        if missing:
            rec["validation_issues"].append("Missing " + ", ".join(missing))
            warning_lines.append(
                "{0}: missing {1}".format(
                    rec["employee_name"] or employee_id,
                    ", ".join(missing),
                )
            )
        if rec["raw_employee_name"] and rec["raw_employee_name"] != rec["employee_name"]:
            warning_lines.append(
                "{0}: employee name was cleaned for WPS sheet as '{1}'".format(
                    rec["raw_employee_name"],
                    rec["employee_name"] or employee_id,
                )
            )
        if rec["validation_issues"]:
            warning_lines.append(
                "{0}: validation issue(s): {1}".format(
                    rec["employee_name"] or employee_id,
                    "; ".join(rec["validation_issues"]),
                )
            )
        if len(rec["salary_slips"]) > 1:
            warning_lines.append(
                "{0}: combined {1} Salary Slips ({2})".format(
                    rec["employee_name"] or employee_id,
                    len(rec["salary_slips"]),
                    ", ".join(rec["salary_slips"]),
                )
            )
        rows.append(rec)

    return rows, warning_lines


def _get_salary_detail_totals(slip_names):
    totals = defaultdict(lambda: {"earnings": defaultdict(float), "deductions": defaultdict(float)})
    for chunk in _chunks(slip_names, 500):
        for row in frappe.get_all(
            "Salary Detail",
            filters=[["parent", "in", chunk]],
            fields=["parent", "parentfield", "salary_component", "amount"],
        ):
            parentfield = row.get("parentfield")
            if parentfield not in ("earnings", "deductions"):
                continue
            key = _norm(row.get("salary_component"))
            if not key:
                continue
            totals[row.parent][parentfield][key] += flt(row.get("amount"))
    return totals


def _get_employee_map(employee_names):
    employees = {}
    fields = _safe_fields(
        "Employee",
        [
            "name",
            "employee_name",
            "department",
            "bank_name",
            "bank_ac_no",
            "bank_account_no",
            "iban",
            "employee_number",
            "iqama_national_id",
            "passport_number",
            "custom_employee_number",
            "custom_iqama_national_id",
            "custom_iqama_number",
            "custom_id_number",
            "custom_iban",
        ] + _employee_iban_fieldnames(),
    )
    for chunk in _chunks(sorted(set(employee_names)), 500):
        for emp in frappe.get_all("Employee", filters=[["name", "in", chunk]], fields=fields):
            employees[emp.name] = emp
    return employees


def _build_workbook(rows):
    wb = load_workbook(_template_path(), keep_vba=False, data_only=False)
    ws = wb[SHEET_NAME]

    for offset, row in enumerate(rows):
        row_idx = DATA_START_ROW + offset
        ws.cell(row=row_idx, column=1, value=row["bank_name"])
        ws.cell(row=row_idx, column=2, value=row["iban"])
        ws.cell(row=row_idx, column=3, value=row["employee_name"])
        ws.cell(row=row_idx, column=4, value=row["employee_number"])
        ws.cell(row=row_idx, column=5, value=row["national_id"])
        ws.cell(row=row_idx, column=6, value=f"=G{row_idx}+H{row_idx}+I{row_idx}-J{row_idx}")
        ws.cell(row=row_idx, column=7, value=row["basic"])
        ws.cell(row=row_idx, column=8, value=row["housing"])
        ws.cell(row=row_idx, column=9, value=row["other_earnings"])
        ws.cell(row=row_idx, column=10, value=row["deductions"])
        ws.cell(row=row_idx, column=11).value = None
        ws.cell(row=row_idx, column=12).value = None
        ws.cell(row=row_idx, column=13, value="; ".join(row.get("validation_issues") or []))
        ws.cell(row=row_idx, column=14, value=row["department"])
        if row.get("validation_issues"):
            fill = IQAMA_VALIDATION_FILL if row.get("has_iqama_issue") else VALIDATION_FILL
            for col_idx in range(1, 15):
                ws.cell(row=row_idx, column=col_idx).fill = fill

    # Keep the row after the generated data blank while preserving its formula.
    first_blank_row = DATA_START_ROW + len(rows)
    for col in list(range(1, 6)) + list(range(7, 15)):
        ws.cell(row=first_blank_row, column=col).value = None
    ws.cell(
        row=first_blank_row,
        column=6,
        value=f"=G{first_blank_row}+H{first_blank_row}+I{first_blank_row}-J{first_blank_row}",
    )

    try:
        wb.calculation.fullCalcOnLoad = True
        wb.calculation.forceFullCalc = True
    except Exception:
        pass

    return wb


def _save_workbook_and_link(wb, doctype, docname, fieldname, filename, label):
    buf = io.BytesIO()
    wb.save(buf)
    content_bytes = buf.getvalue()
    buf.close()
    wb.close()

    safe_name = _safe_filename(filename)
    private_dir = frappe.get_site_path("private", "files")
    os.makedirs(private_dir, exist_ok=True)

    final_name = _unique_filename(private_dir, safe_name)
    full_path = os.path.join(private_dir, final_name)
    with open(full_path, "wb") as f:
        f.write(content_bytes)

    file_url = f"/private/files/{final_name}"

    try:
        file_doc = frappe.get_doc({
            "doctype": "File",
            "file_name": final_name,
            "file_url": file_url,
            "is_private": 1,
            "folder": "Home/Attachments",
            "attached_to_doctype": doctype,
            "attached_to_name": docname,
            "attached_to_field": fieldname,
        })
        file_doc.flags.ignore_permissions = True
        file_doc.flags.ignore_mandatory = True
        file_doc.insert(ignore_permissions=True)
        frappe.db.commit()
    except Exception:
        frappe.log_error(
            title=f"Payroll WPS generate: File record creation FAILED for {label}",
            message=frappe.get_traceback(),
        )

    frappe.db.set_value(doctype, docname, fieldname, file_url, update_modified=False)
    frappe.db.commit()
    return file_url


def _safe_fields(doctype, fieldnames):
    meta = frappe.get_meta(doctype)
    out = []
    for fieldname in _dedupe(fieldnames):
        if fieldname == "name" or meta.has_field(fieldname):
            out.append(fieldname)
    return out


def _require_field(doctype, fieldname):
    if not frappe.get_meta(doctype).has_field(fieldname):
        frappe.throw(_("{0}.{1} is required for WPS export").format(doctype, fieldname))


def _component_keys(primary, aliases):
    names = [primary] if primary else []
    names.extend(aliases)
    return {_norm(name) for name in names if _norm(name)}


def _pick_component_total(components, wanted_keys):
    return round(sum(amount for key, amount in components.items() if key in wanted_keys), 2)


def _has_component(components, wanted_keys):
    return any(key in wanted_keys for key in components)


def _deduction_amount(slip, gross_pay, deductions_by_component):
    """Use Net Pay as the source of truth so loan repayments are included."""
    net_pay = slip.get("net_pay")
    if net_pay not in (None, ""):
        return round(flt(gross_pay) - flt(net_pay), 2)

    total_deduction = slip.get("total_deduction")
    if total_deduction not in (None, ""):
        return round(flt(total_deduction), 2)

    return round(sum(deductions_by_component.values()), 2)


def _is_zero_net_pay(slip):
    net_pay = slip.get("net_pay")
    if net_pay in (None, ""):
        return False
    return abs(flt(net_pay)) < 0.005


def _bank_name(slip, emp):
    raw_bank = _first_value(slip, ("bank_name",)) or _first_value(emp, ("bank_name",))
    iban = _iban(slip, emp)
    return bank_display_for(raw_bank, iban) or raw_bank or ""


def _iban(slip, emp):
    value = (
        _first_value(emp, _employee_iban_fieldnames())
        or _first_value(emp, ("iban", "custom_iban", "bank_ac_no", "bank_account_no"))
        or _first_value(slip, ("bank_account_no", "bank_ac_no"))
    )
    return _clean_iban(value)


def _employee_iban_fieldnames():
    exact = []
    bank_account = []
    try:
        fields = frappe.get_meta("Employee").fields
    except Exception:
        return []
    for df in fields:
        fieldname = getattr(df, "fieldname", None)
        if not fieldname:
            continue
        label = getattr(df, "label", None) or ""
        key = _field_key(f"{fieldname} {label}")
        if "iban" in key:
            exact.append(fieldname)
        elif "bank" in key and ("account" in key or "acno" in key):
            bank_account.append(fieldname)
    return _dedupe(exact + bank_account)


def _employee_number(slip, emp, fallback):
    for value in (
        _first_value(slip, ("employee_no", "employee_number")),
        _first_value(emp, ("employee_number", "custom_employee_number")),
        fallback,
    ):
        cleaned = _clean_employee_number(value)
        if cleaned:
            return cleaned
    return ""


def _national_id(emp):
    return _clean_digits(_first_value(
        emp,
        (
            "iqama_national_id",
            "custom_iqama_national_id",
            "custom_iqama_number",
            "custom_id_number",
            "passport_number",
        ),
    ))


def _department_value(slip, emp):
    department = slip.get("department") or emp.get("department")
    if not department:
        return ""
    return frappe.db.get_value("Department", department, "department_name") or department


def _first_value(source, fieldnames):
    for fieldname in fieldnames:
        value = source.get(fieldname) if source else None
        if value not in (None, ""):
            return _clean_text(value)
    return ""


def _clean_text(value):
    if value is None:
        return ""
    return str(value).strip()


def _clean_iban(value):
    return "".join(char for char in _clean_text(value) if char.isalnum()).upper()


def _clean_digits(value):
    return "".join(char for char in _clean_text(value) if char.isdigit())


def _clean_employee_name(value):
    cleaned = "".join(
        char if char.isascii() and (char.isalnum() or char.isspace()) else " "
        for char in _clean_text(value)
    )
    return " ".join(cleaned.split())


def _clean_employee_number(value):
    text = _clean_text(value)
    if not text:
        return ""
    digits = "".join(char for char in text if char.isdigit())
    if not digits:
        return ""
    return str(int(digits))


def _component_names_for_log(components):
    if not components:
        return "none"
    return ", ".join(sorted(components.keys()))


def _iban_error(iban):
    if not iban:
        return "missing IBAN on Employee"
    if len(iban) != 24:
        return "IBAN must be 24 characters without spaces"
    if not iban.startswith("SA"):
        return "Saudi IBAN must start with SA"
    if not iban[2:4].isdigit():
        return "IBAN check digits must be numeric"
    if not iban[4:6].isdigit():
        return "Saudi IBAN bank code must be numeric"
    if not iban[6:].isalnum():
        return "Saudi IBAN account number must contain letters/numbers only"
    if not _passes_iban_checksum(iban):
        return "IBAN checksum is invalid"
    return ""


def _iqama_error(national_id):
    if not national_id:
        return "missing Iqama/National ID"
    if len(national_id) != 10:
        return "Iqama/National ID must be 10 digits"
    return ""


def _passes_iban_checksum(iban):
    rearranged = iban[4:] + iban[:4]
    remainder = 0
    for char in rearranged:
        if char.isdigit():
            digits = char
        elif char.isalpha():
            digits = str(ord(char) - 55)
        else:
            return False
        for digit in digits:
            remainder = (remainder * 10 + int(digit)) % 97
    return remainder == 1


def _project_label(project):
    return frappe.db.get_value("Project", project, "project_name") or project or "Project"


def _department_label(department):
    if not department:
        return ""
    return frappe.db.get_value("Department", department, "department_name") or department


def _work_location_label(work_location):
    if not work_location:
        return ""
    try:
        meta = frappe.get_meta("Work Location")
        for fieldname in ("location_name", "work_location_name", "title"):
            if meta.has_field(fieldname):
                return frappe.db.get_value("Work Location", work_location, fieldname) or work_location
    except Exception:
        pass
    return work_location


def _period_label(start_date, end_date):
    start = getdate(start_date)
    end = getdate(end_date)
    if start.year == end.year and start.month == end.month:
        return start.strftime("%B %Y")
    return f"{start.isoformat()} to {end.isoformat()}"


def _template_path():
    return frappe.get_app_path("erc_payroll_automation", "templates", TEMPLATE_FILENAME)


def _safe_filename(name):
    cleaned = re.sub(r"[^\w\s\-.()]+", "_", name, flags=re.UNICODE)
    cleaned = re.sub(r"\s+", " ", cleaned).strip()
    return cleaned or "wps_payroll_upload.xlsm"


def _unique_filename(directory, name):
    if not os.path.exists(os.path.join(directory, name)):
        return name
    base, ext = os.path.splitext(name)
    counter = 1
    while True:
        candidate = f"{base}_{counter}{ext}"
        if not os.path.exists(os.path.join(directory, candidate)):
            return candidate
        counter += 1


def _chunks(values, size):
    values = list(values or [])
    for idx in range(0, len(values), size):
        yield values[idx:idx + size]


def _norm(value):
    return " ".join(str(value or "").strip().lower().split())


def _field_key(value):
    return "".join(char for char in str(value or "").strip().lower() if char.isalnum())


def _dedupe(values):
    out = []
    seen = set()
    for value in values or []:
        if not value or value in seen:
            continue
        seen.add(value)
        out.append(value)
    return out
