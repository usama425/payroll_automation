"""Build the Additional Salary import Excel ("Additional_Salary_RSG.xlsx").

This file is designed to be uploaded into ERPNext via Data Import → Additional Salary.
Format matches the historical example exactly:

    Employee | Payroll Date | Salary Component | Amount | Overwrite Salary Structure Amount

Mapping rule (confirmed against the Feb-2026 example file — 15/16 auto-generated rows
matched the source columns 1:1):

    Customer file column → ERPNext Salary Component
    --------------------------------------------------
    Other Income (non-zero)   → "Other Additions"
    Overtime (non-zero)       → "Overtime"
    Deductions (non-zero)     → "Deductions"
    HQ Ded (non-zero)         → "HQ Deductions"

One row per (Employee, Salary Component) with non-zero amount.
Rows with all four fields = 0 produce no Additional Salary entries.
"""

import frappe
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from .file_attach import save_excel_and_link


# Source column on Payroll Import Row → Salary Component name in ERPNext
SOURCE_TO_COMPONENT = [
    ("other_income",  "Other Additions"),
    ("overtime",      "Overtime"),
    ("deductions",    "Deductions"),
    ("hq_deductions", "HQ Deductions"),
]


def generate(run, template) -> str:
    """Build the Additional Salary file. Returns the file URL."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Additional Salary"

    # Row 1: headers (exact match to ERPNext Data Import header expectations)
    headers = [
        "Employee", "Payroll Date", "Salary Component",
        "Amount", "Overwrite Salary Structure Amount",
    ]
    for col_idx, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col_idx, value=h)
        c.font = Font(bold=True)
        c.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        c.alignment = Alignment(horizontal="center")

    payroll_date = run.posting_date or run.payroll_period_end

    excel_row = 2
    for row in run.parsed_rows:
        if not row.employee:
            continue
        for fieldname, component in SOURCE_TO_COMPONENT:
            amount = getattr(row, fieldname, None)
            if amount in (None, 0, 0.0):
                continue
            try:
                amount = float(amount)
            except (TypeError, ValueError):
                continue
            if abs(amount) < 0.01:
                continue
            ws.cell(row=excel_row, column=1, value=row.employee)
            ws.cell(row=excel_row, column=2, value=payroll_date)
            ws.cell(row=excel_row, column=3, value=component)
            ws.cell(row=excel_row, column=4, value=amount).number_format = "#,##0.00"
            ws.cell(row=excel_row, column=5, value=1)
            excel_row += 1

    # Column widths
    widths = {1: 18, 2: 14, 3: 22, 4: 14, 5: 32}
    for col, w in widths.items():
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = w

    filename = f"Additional_Salary_RSG_{run.name}.xlsx"
    return save_excel_and_link(
        wb, run.name, "additional_salary_file", filename, "Additional Salary",
    )
